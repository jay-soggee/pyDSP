import numpy as np
import pyaudio
import sys
import keyboard
import random
import openpyxl  as  op

RATE = 48000
CHUNK = int(RATE/10)
HEAD = 10
DIS = 1
Vs = 339
DEG_FIVE = np.pi*float(5)/180

max_delay = int(RATE*2*HEAD*.01/Vs)
aheadL = np.zeros(max_delay, dtype=np.int16)
aheadR = np.zeros(max_delay, dtype=np.int16)
out = np.zeros(CHUNK*2, dtype=np.int16)

def Get_delay(dir_angle):
    distance = 2 * HEAD * 0.01 * np.abs(np.sin(dir_angle))
    delay = int(distance * RATE/Vs)
    
    return distance, delay

def Sound_rendering(signal, dir):

    distance, delay = Get_delay(dir)

    if dir >= 0:
        for i in range(CHUNK):
            # ITD
            if i < delay:
               out[i*2] = aheadL[max_delay - delay + i]
            else:
               out[i*2] = signal[(i-delay)*2]
            # IID
            out[i*2] = int(out[i*2]*(DIS/(DIS+distance)))
            out[i*2+1] = signal[i*2+1]
    else:
        for i in range(CHUNK):
            # ITD
            if i < delay:
               out[i*2+1] = aheadR[max_delay - delay + i]
            else:
               out[i*2+1] = signal[(i-delay)*2]
            # IID
            out[i*2+1] = int(out[i*2+1]*(DIS/(DIS+distance)))
            out[i*2] = signal[i*2]
    for i in range(max_delay):
        aheadL[i] = signal[(CHUNK - max_delay + i)*2]
        aheadR[i] = signal[(CHUNK - max_delay + i)*2]

    return out


p=pyaudio.PyAudio()
stream = p.open(format=pyaudio.paInt16, 
                channels=2, 
                rate=RATE, 
                input=True, 
                output=True,
                frames_per_buffer=CHUNK, 
                input_device_index=0
                )

rendering = True
deg = float(sys.argv[1])
target_dir = np.pi*deg/180
dir = target_dir

## GAME

GAME_IDLE   = 0
GAME_RUN    = 1
GAME_QUEST  = 2
game_state = 0
usr_guess = 0


## DATA
COL_ANS     = 1
COL_GUSS    = 2
xlsx_curr_row   = 1
wb = op.Workbook()
ws = wb.create_sheet("game2")

while(True):
    samples = stream.read(CHUNK)
    in_data = np.fromstring(samples, dtype=np.int16)
    out = Sound_rendering(in_data, dir)
    y = out.tostring()
    stream.write(y)
    if keyboard.is_pressed('q'):
        sys.stdin.flush()
        wb.save(r"result.xlsx")
        break
    elif keyboard.is_pressed('s'):
        if rendering:
            rendering = False
            dir = 0
            print("Sound rendering off.")
        else:
            rendering = True
            dir = target_dir
            print("Sound rendering on.")
    elif keyboard.is_pressed('u') & (game_state == GAME_IDLE):
        if (deg + 5 > 90): deg = 90 
        else : deg = deg + 5
        print("now deg = "+ f"{deg}")
        target_dir = np.pi*deg/180
        dir = target_dir
    elif keyboard.is_pressed('d') & (game_state == GAME_IDLE):
        if (deg - 5 < -90): deg = -90 
        else : deg = deg - 5
        print("now deg = "+ f"{deg}")
        target_dir = np.pi*deg/180
        dir = target_dir
    elif keyboard.is_pressed('g') & (game_state == GAME_IDLE):
        print("game mode on.")
        game_state = GAME_RUN
        deg = 0
        target_dir = np.pi*deg/180
        dir = target_dir
    
    if (game_state == GAME_RUN):
        game_state = GAME_QUEST
        deg = random.randrange(-18, 19)*5
        target_dir = np.pi*deg/180
        dir = target_dir
        print("guess degree of the sound : ")
        print("    now you guessing : "+f"{usr_guess}")
    elif (game_state == GAME_QUEST):
        if keyboard.is_pressed('u'):
            if (usr_guess + 5 > 90): usr_guess = 90 
            else: usr_guess = usr_guess + 5
            print("    now you guessing : "+ f"{usr_guess}")
        elif keyboard.is_pressed('d'):
            if (usr_guess - 5 < -90): usr_guess = -90 
            else: usr_guess = usr_guess - 5
            print("    now you guessing : "+ f"{usr_guess}")
        elif keyboard.is_pressed('enter'):
            print("    you guessed : "+ f"{usr_guess}")
            #print("    and the answer was : "+ f"{deg}")
            ws.cell(row=xlsx_curr_row, column=COL_ANS).value = deg
            ws.cell(row=xlsx_curr_row, column=COL_GUSS).value = usr_guess
            xlsx_curr_row = xlsx_curr_row + 1
            print("    trial-"+ f"{xlsx_curr_row}")
            game_state = GAME_RUN
            

    
        
      
stream.stop_stream()
stream.close()

p.terminate()

